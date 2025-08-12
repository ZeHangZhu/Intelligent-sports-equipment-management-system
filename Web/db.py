import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

'''初始化数据库'''
dbEquipment = sqlite3.connect('./data/Equipment.db',check_same_thread = False)
comEquipment = dbEquipment.cursor()

'''数据库操作函数'''
# 增加器材
def addEquipment(type,code):
    # 获取当前最大ID
    comEquipment.execute('SELECT MAX(ID) FROM EQUIPMENT')
    max_id = comEquipment.fetchone()[0]
    if max_id is None:
        max_id = 0
    new_id = max_id + 1
    
    # 插入新记录
    comEquipment.execute(f'''INSERT INTO EQUIPMENT (ID, CODE, TYPE, STATE) 
                         VALUES ({new_id}, {code},"{type}",0);''')
    dbEquipment.commit()

# 增加日志
def addLog(name,GradeAndClass,code):
    # 获取当前最大ID
    comEquipment.execute('SELECT MAX(ID) FROM LOG')
    max_id = comEquipment.fetchone()[0]
    if max_id is None:
        max_id = 0
    new_id = max_id + 1

    # 获取当前时间
    now = datetime.now()

    # 提取年、月、日、小时和分钟
    year = now.year
    month = now.month
    day = now.day
    hour = now.hour
    minute = now.minute

    # print(f"当前时间是: {year}年{month}月{day}日 {hour}点{minute}分")
    DateAndTime =f"{year}年{month}月{day}日 {hour}点{minute}分"

    # 插入新记录
    comEquipment.execute(f'''INSERT INTO LOG (ID, NAME, CLASS, CODE, TIME, STATE) 
                         VALUES ({new_id}, "{name}","{GradeAndClass}",{code},"{DateAndTime}",0)''')
    dbEquipment.commit()

# 重新建立数据库文件并创建表在数据库中
# 慎重使用,该功能会使数据丢失
def reInstallEquipment():
    # os.remove("./Data/Equipment.db")
    try:
        comEquipment.execute('''CREATE TABLE EQUIPMENT
        (ID INT PRIMARY KEY      NOT NULL,
            CODE           INT     NOT NULL,
            TYPE           TEXT     NOT NULL,
            STATE          INT      NOT NULL);''')
        comEquipment.execute('''CREATE TABLE LOG
        (ID INT PRIMARY KEY      NOT NULL,
            NAME           TEXT     NOT NULL,
            CLASS          TEXT     NOT NULL,
            CODE           INT     NOT NULL,
            TIME           TEXT     NOT NULL,
            STATE          INT      NOT NULL,
            TIME2          TEXT);''')
    except Exception as E:
        print(f"[!]{E}")

# 用于生成code
def generate_next_code():
    # 获取当前最大CODE
    comEquipment.execute('SELECT MAX(CODE) FROM EQUIPMENT')
    max_code = comEquipment.fetchone()[0]
    if max_code is None:
        max_code = 0
    next_code = max_code + 1
    return next_code

# 用于检查是否存在CODE
def check_code_exists(code):
    # 查询CODE是否存在
    query = "SELECT COUNT(*) FROM EQUIPMENT WHERE CODE = ?"
    comEquipment.execute(query, (code,))
    
    # 获取查询结果
    result = comEquipment.fetchone()
    count = result[0]
    
    # 返回查询结果
    return count > 0
def update_equipment_state(code, state):
    """
    更新EQUIPMENT表中指定CODE的STATE字段
    :param code: int, 设备的CODE
    :param state: int, 要更新的STATE值
    """
    try:
        # 更新STATE字段
        comEquipment.execute('''
        UPDATE EQUIPMENT
        SET STATE = ?
        WHERE CODE = ?
        ''', (state, code))
        
        # 提交事务
        dbEquipment.commit()
        print(f"Successfully updated the state for code {code} to {state}")
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")

def get_equipment_state(code):
    """
    获取EQUIPMENT表中指定CODE的STATE字段的值
    :param code: int, 设备的CODE
    :return: int, 对应的STATE值
    """
    try:
        # 查询STATE字段
        comEquipment.execute('''
        SELECT STATE FROM EQUIPMENT
        WHERE CODE = ?
        ''', (code,))
        
        # 获取查询结果
        result = comEquipment.fetchone()
        if result:
            return result[0]
        else:
            print(f"No record found for code {code}")
            return -25565   # 返回错误代码
    except sqlite3.Error as e:
        print(f"An error occurred: {e}")

def get_log_state(code):
    query = """
    SELECT STATE FROM LOG WHERE CODE = ? ORDER BY ID DESC LIMIT 1
    """
    comEquipment.execute(query, (code,))
    result = comEquipment.fetchone()
    
    if result:
        return result[0]
    else:
        return -25565

def update_log_state(code):
    # 找到最后一条对应 code 的记录
    query = """
    SELECT ID FROM LOG WHERE CODE = ? ORDER BY ID DESC LIMIT 1
    """
    comEquipment.execute(query, (code,))
    result = comEquipment.fetchone()
    
    if result:
        try:
            last_id = result[0]
            
            # 更新该记录的 STATE 为 1
            update_query = """
            UPDATE LOG SET STATE = 1 WHERE ID = ?
            """
            comEquipment.execute(update_query, (last_id,))
            dbEquipment.commit()
            print(f"Updated the state of record with ID {last_id} to 1.")
        except Exception as e:
            print(f"Error updating state: {e}")

    else:
        print(f"No record found for code {code}.")

def delete_equipment_by_code(code):
    """
    删除数据库中CODE对应的设备记录

    参数:
    db_path (str): SQLite数据库文件的路径
    code (int): 要删除的设备CODE
    """
    try:
        # 执行删除操作
        sql_delete_query = '''DELETE FROM EQUIPMENT WHERE CODE = ?'''
        comEquipment.execute(sql_delete_query, (code,))
        
        # 提交事务
        dbEquipment.commit()
        print(f"成功删除CODE为{code}的设备记录")
    
    except Exception as E:
        print(f"错误:{E}")

def get_equipment_type_by_code(code):
    """
    根据设备的CODE获取对应的TYPE

    参数:
    code (int): 设备的CODE

    返回:
    str: 设备的TYPE，如果未找到则返回None
    """
    try:
        # 执行查询操作
        sql_select_query = '''SELECT TYPE FROM EQUIPMENT WHERE CODE = ?'''
        comEquipment.execute(sql_select_query, (code,))
        
        # 获取查询结果
        result = comEquipment.fetchone()
        if result:
            return result[0]
        else:
            print(f"未找到CODE为{code}的设备记录")
            return None
    
    except Exception as E:
        print(f"错误:{E}")

def add_return_time(code):
    """
    在最近的与指定 CODE 匹配的 LOG 记录中添加 TIME2 数据为当前时间

    :param code: int, 设备的 CODE
    """
    try:
        # 获取当前时间
        now = datetime.now()
        # 格式化时间为所需的字符串格式
        current_time = now.strftime("%Y年%m月%d日 %H点%M分")

        # 找到最后一条对应 code 的记录
        query = """
        SELECT ID FROM LOG WHERE CODE = ? ORDER BY ID DESC LIMIT 1
        """
        comEquipment.execute(query, (code,))
        result = comEquipment.fetchone()

        if result:
            last_id = result[0]

            # 更新该记录的 TIME2 字段为当前时间
            update_query = """
            UPDATE LOG SET TIME2 = ? WHERE ID = ?
            """
            comEquipment.execute(update_query, (current_time, last_id))
            dbEquipment.commit()
            print(f"成功更新 ID 为 {last_id} 的记录的 TIME2 字段为 {current_time}")
            return 1
        else:
            print(f"未找到 CODE 为 {code} 的记录")
            return 0

    except Exception as e:
        return -1
def export_log_table_to_excel(filename):
    """
    将数据库中 log 表的所有数据导出为 Excel 表格。

    :param filename: 导出的 Excel 文件的文件名，需包含文件扩展名（如 .xlsx）
    """
    try:
        # 查询 log 表中的所有数据
        query = "SELECT * FROM LOG"
        comEquipment.execute(query)
        data = comEquipment.fetchall()

        # 获取列名
        column_names = [description[0] for description in comEquipment.description]

        # 定义英文表头到中文表头的映射
        header_mapping = {
            "ID": "编号",
            "NAME": "姓名",
            "CLASS": "班级",
            "CODE": "编号",
            "TIME": "借用时间",
            "STATE": "状态",
            "TIME2": "归还时间"
        }

        # 将英文表头转换为中文表头
        chinese_column_names = [header_mapping.get(col, col) for col in column_names]

        # 将数据转换为 DataFrame
        df = pd.DataFrame(data, columns=chinese_column_names)

        # 将 DataFrame 保存为 Excel 文件
        df.to_excel(filename, index=False, engine='openpyxl')

        # 加载保存的 Excel 文件
        wb = load_workbook(filename)
        ws = wb.active

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存修改后的 Excel 文件
        wb.save(filename)

        print(f"成功将 log 表的数据导出到 {filename}")

    except Exception as e:
        print(f"导出数据时出现错误: {e}")

